"""
文件预处理器 - 检查文件大小、行数、列数等，优化性能
"""
import os
from pathlib import Path
from typing import Tuple, Dict, Optional
import openpyxl
from openpyxl import load_workbook
try:
    import pyxlsb
except ImportError:
    pyxlsb = None

from .models import FileFormat
from .exceptions import FileReadError
from .logger import DualLogger
from .models import LogLevel


class FilePreprocessor:
    """文件预处理器"""
    
    def __init__(self, config, logger: Optional[DualLogger] = None):
        self.config = config
        self.logger = logger
    
    def preprocess_file(self, file_path: str, format: FileFormat) -> Dict:
        """
        预处理文件：检查大小、行数、列数等
        返回预处理信息字典
        """
        info = {
            "file_path": file_path,
            "file_size_mb": 0.0,
            "estimated_rows": 0,
            "estimated_cols": 0,
            "warnings": [],
            "should_optimize": False,
        }
        
        # 1. 检查文件大小
        file_size = os.path.getsize(file_path)
        file_size_mb = file_size / (1024 * 1024)
        info["file_size_mb"] = file_size_mb
        
        if file_size_mb > self.config.max_file_size_mb:
            warning = f"文件大小 {file_size_mb:.2f}MB 超过建议值 {self.config.max_file_size_mb}MB，处理可能较慢"
            info["warnings"].append(warning)
            info["should_optimize"] = True
            if self.logger:
                self.logger.log("preprocess.warning", level=LogLevel.WARN, message=warning)
        
        # 2. 快速检查行数和列数（不加载完整数据）
        if format == FileFormat.xlsx:
            try:
                # 使用只读模式快速检查
                wb = load_workbook(file_path, read_only=True, data_only=True)
                if wb.sheetnames:
                    # 检查第一个 sheet 的尺寸（作为参考）
                    ws = wb[wb.sheetnames[0]]
                    max_row = ws.max_row or 0
                    max_col = ws.max_column or 0
                    info["estimated_rows"] = max_row
                    info["estimated_cols"] = max_col
                    
                    if max_row > self.config.max_rows:
                        warning = f"Sheet 行数 {max_row} 超过限制 {self.config.max_rows}，将截断处理"
                        info["warnings"].append(warning)
                        info["should_optimize"] = True
                        if self.logger:
                            self.logger.log("preprocess.warning", level=LogLevel.WARN, message=warning,
                                          metrics={"rows": max_row, "limit": self.config.max_rows})
                    
                    if max_col > self.config.max_cols:
                        warning = f"Sheet 列数 {max_col} 超过限制 {self.config.max_cols}，将截断处理"
                        info["warnings"].append(warning)
                        info["should_optimize"] = True
                        if self.logger:
                            self.logger.log("preprocess.warning", level=LogLevel.WARN, message=warning,
                                          metrics={"cols": max_col, "limit": self.config.max_cols})
                
                wb.close()
            except Exception as e:
                # 如果只读模式失败，记录但不阻止处理
                if self.logger:
                    self.logger.log("preprocess.skip", level=LogLevel.WARN,
                                  message=f"无法快速检查文件尺寸: {str(e)}")
        
        elif format == FileFormat.xlsb:
            # xlsb 格式的快速检查
            if pyxlsb:
                try:
                    with pyxlsb.open_workbook(file_path) as wb:
                        if wb.sheets:
                            sheet_name = list(wb.sheets)[0]
                            ws = wb.get_sheet(sheet_name)
                            # xlsb 没有直接获取 max_row 的方法，需要估算
                            # 这里简化处理
                            pass
                except Exception:
                    pass
        
        # 3. 记录预处理信息
        if self.logger:
            self.logger.log("preprocess.complete", 
                          metrics={
                              "file_size_mb": round(file_size_mb, 2),
                              "estimated_rows": info["estimated_rows"],
                              "estimated_cols": info["estimated_cols"],
                              "warnings_count": len(info["warnings"]),
                              "should_optimize": info["should_optimize"]
                          })
        
        return info
    
    def get_optimized_limits(self, estimated_rows: int, estimated_cols: int) -> Tuple[int, int]:
        """
        根据估算的行列数，返回优化的读取限制
        """
        max_row = min(estimated_rows, self.config.max_rows) if estimated_rows > 0 else self.config.max_rows
        max_col = min(estimated_cols, self.config.max_cols) if estimated_cols > 0 else self.config.max_cols
        return max_row, max_col

