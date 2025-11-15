"""
文件读取器 - 支持 xlsx/xlsb/csv
"""
import pandas as pd
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
import openpyxl
from openpyxl import load_workbook
try:
    import pyxlsb
except ImportError:
    pyxlsb = None

from .models import FileFormat
from .exceptions import UnsupportedFormatError, FileReadError, InvalidArgumentError


def detect_format(file_path: str) -> FileFormat:
    """检测文件格式"""
    ext = Path(file_path).suffix.lower()
    if ext == ".xlsx":
        return FileFormat.xlsx
    elif ext == ".xlsb":
        return FileFormat.xlsb
    elif ext == ".csv":
        return FileFormat.csv
    else:
        raise UnsupportedFormatError(f"Unsupported file extension: {ext}")


def read_xlsx_sheet(file_path: str, sheet_name: str, include_hidden: bool = False, 
                   max_rows: Optional[int] = None, max_cols: Optional[int] = None,
                   enable_style_scan: bool = True, enable_border_scan: bool = True,
                   style_scan_limit: Optional[int] = None, border_scan_limit: Optional[int] = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    读取 xlsx 文件的指定 sheet（单个 sheet）
    返回: (DataFrame, metadata_dict)
    metadata 包含: borders, styles, merged_cells, hidden_rows, hidden_cols
    """
    try:
        # 使用普通模式（read_only=False）以确保能访问所有属性
        # 虽然速度稍慢，但能保证功能完整性
        wb = load_workbook(file_path, data_only=True, read_only=False)
        
        if sheet_name not in wb.sheetnames:
            wb.close()
            raise FileReadError(f"Sheet '{sheet_name}' not found in file")
        
        ws = wb[sheet_name]
        
        # 读取数据（应用限制）
        data = []
        original_max_row = ws.max_row or 0
        original_max_col = ws.max_column or 0
        
        # 应用行数和列数限制
        max_row = min(original_max_row, max_rows) if max_rows else original_max_row
        max_col = min(original_max_col, max_cols) if max_cols else original_max_col
        
        # 使用 values_only=True 提升性能（只读取值，不读取样式）
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True), start=1):
            # values_only=True 时，row 直接是值列表
            row_data = list(row)
            data.append(row_data)
        
        # 转换为 DataFrame
        if data:
            # 确保所有行长度一致
            max_len = max(len(row) for row in data) if data else 0
            data = [row + [None] * (max_len - len(row)) for row in data]
            df = pd.DataFrame(data)
        else:
            df = pd.DataFrame()
        
        # 收集元数据
        metadata = {
            "borders": {},
            "styles": {},
            "merged_cells": [],
            "hidden_rows": set(),
            "hidden_cols": set(),
        }
        
        # 合并单元格（普通模式下可以直接访问）
        try:
            if hasattr(ws, 'merged_cells') and hasattr(ws.merged_cells, 'ranges'):
                for merged_range in ws.merged_cells.ranges:
                    metadata["merged_cells"].append({
                        "min_row": merged_range.min_row - 1,  # 转为0-based
                        "max_row": merged_range.max_row - 1,
                        "min_col": merged_range.min_col - 1,
                        "max_col": merged_range.max_col - 1,
                    })
        except (AttributeError, TypeError) as e:
            # 如果无法获取合并单元格信息，继续执行
            pass
        
        # 隐藏行列
        if not include_hidden:
            for row_idx, row in enumerate(ws.iter_rows(), start=1):
                if row and hasattr(row[0], 'row') and ws.row_dimensions[row[0].row].hidden:
                    metadata["hidden_rows"].add(row_idx - 1)  # 0-based
            
            for col_idx in range(1, max_col + 1):
                if ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden:
                    metadata["hidden_cols"].add(col_idx - 1)  # 0-based
        
        # 边框和样式（仅在启用时扫描，且限制扫描范围以提升性能）
        if enable_border_scan or enable_style_scan:
            scan_limit = max(
                border_scan_limit if border_scan_limit else 0,
                style_scan_limit if style_scan_limit else 0
            ) or max_row
            
            scan_rows = min(scan_limit, max_row)
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_rows, min_col=1, max_col=max_col, values_only=False), start=1):
                for col_idx, cell in enumerate(row, start=1):
                    key = (row_idx - 1, col_idx - 1)  # 0-based
                    
                    # 边框（仅在启用时扫描）
                    if enable_border_scan:
                        if cell.border and cell.border.left.style:
                            if key not in metadata["borders"]:
                                metadata["borders"][key] = {}
                            metadata["borders"][key]["left"] = cell.border.left.style is not None
                        if cell.border and cell.border.right.style:
                            if key not in metadata["borders"]:
                                metadata["borders"][key] = {}
                            metadata["borders"][key]["right"] = cell.border.right.style is not None
                        if cell.border and cell.border.top.style:
                            if key not in metadata["borders"]:
                                metadata["borders"][key] = {}
                            metadata["borders"][key]["top"] = cell.border.top.style is not None
                        if cell.border and cell.border.bottom.style:
                            if key not in metadata["borders"]:
                                metadata["borders"][key] = {}
                            metadata["borders"][key]["bottom"] = cell.border.bottom.style is not None
                    
                    # 样式（仅在启用时扫描）
                    if enable_style_scan:
                        if cell.font and cell.font.bold:
                            if key not in metadata["styles"]:
                                metadata["styles"][key] = {}
                            metadata["styles"][key]["bold"] = True
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            if key not in metadata["styles"]:
                                metadata["styles"][key] = {}
                            metadata["styles"][key]["bg_color"] = str(cell.fill.start_color.rgb)
        
        wb.close()
        return df, metadata
    
    except Exception as e:
        raise FileReadError(f"Failed to read xlsx file: {str(e)}")


def read_xlsb_sheet(file_path: str, sheet_name: str, include_hidden: bool = False) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    读取 xlsb 文件的指定 sheet
    """
    if pyxlsb is None:
        raise FileReadError("pyxlsb library not installed. Install with: pip install pyxlsb")
    
    try:
        with pyxlsb.open_workbook(file_path) as wb:
            if sheet_name not in wb.sheets:
                raise FileReadError(f"Sheet '{sheet_name}' not found in file")
            
            ws = wb.get_sheet(sheet_name)
            
            # 读取数据
            data = []
            for row in ws.rows():
                row_data = [item.v for item in row]
                data.append(row_data)
            
            # 转换为 DataFrame
            if data:
                max_len = max(len(row) for row in data) if data else 0
                data = [row + [None] * (max_len - len(row)) for row in data]
                df = pd.DataFrame(data)
            else:
                df = pd.DataFrame()
            
            # xlsb 格式元数据较少
            metadata = {
                "borders": {},
                "styles": {},
                "merged_cells": [],
                "hidden_rows": set(),
                "hidden_cols": set(),
            }
            
            return df, metadata
    
    except Exception as e:
        raise FileReadError(f"Failed to read xlsb file: {str(e)}")


def read_csv_file(file_path: str) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    读取 CSV 文件
    """
    try:
        # 尝试多种编码
        encodings = ["utf-8", "gbk", "gb2312", "latin-1"]
        df = None
        last_error = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv(file_path, encoding=encoding, dtype=str, keep_default_na=False)
                break
            except UnicodeDecodeError as e:
                last_error = e
                continue
        
        if df is None:
            raise FileReadError(f"Failed to decode CSV file with encodings: {encodings}. Last error: {last_error}")
        
        # CSV 没有样式和边框信息
        metadata = {
            "borders": {},
            "styles": {},
            "merged_cells": [],
            "hidden_rows": set(),
            "hidden_cols": set(),
        }
        
        return df, metadata
    
    except Exception as e:
        if isinstance(e, FileReadError):
            raise
        raise FileReadError(f"Failed to read CSV file: {str(e)}")


def read_single_sheet(file_path: str, sheet_name: str, format: FileFormat,
                     include_hidden: bool = False, max_rows: Optional[int] = None, 
                     max_cols: Optional[int] = None, enable_style_scan: bool = True,
                     enable_border_scan: bool = True, style_scan_limit: Optional[int] = None,
                     border_scan_limit: Optional[int] = None) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """
    读取单个 sheet（用于逐个处理，减少内存占用）
    """
    if format == FileFormat.csv:
        if sheet_name != "__csv__":
            raise InvalidArgumentError("CSV files don't have sheet names")
        return read_csv_file(file_path)
    elif format == FileFormat.xlsx:
        return read_xlsx_sheet(
            file_path, sheet_name, include_hidden,
            max_rows=max_rows, max_cols=max_cols,
            enable_style_scan=enable_style_scan,
            enable_border_scan=enable_border_scan,
            style_scan_limit=style_scan_limit,
            border_scan_limit=border_scan_limit
        )
    elif format == FileFormat.xlsb:
        return read_xlsb_sheet(file_path, sheet_name, include_hidden)
    else:
        raise UnsupportedFormatError(f"Unsupported format: {format}")


def read_file(file_path: str, sheet_name: Optional[List[str]], format: Optional[FileFormat] = None, 
              include_hidden: bool = False, max_rows: Optional[int] = None, max_cols: Optional[int] = None,
              enable_style_scan: bool = True, enable_border_scan: bool = True,
              style_scan_limit: Optional[int] = None, border_scan_limit: Optional[int] = None) -> Dict[str, Tuple[pd.DataFrame, Dict[str, Any]]]:
    """
    统一入口：读取文件并返回所有 sheet 的数据和元数据
    注意：此函数会一次性加载所有 sheet，对于大文件建议使用逐个处理模式
    
    返回: {sheet_name: (DataFrame, metadata)}
    """
    if format is None:
        format = detect_format(file_path)
    
    result = {}
    
    if format == FileFormat.csv:
        if sheet_name is not None and len(sheet_name) > 0:
            raise InvalidArgumentError("sheet_name must be None or empty for CSV files")
        df, metadata = read_csv_file(file_path)
        result["__csv__"] = (df, metadata)
    
    elif format in (FileFormat.xlsx, FileFormat.xlsb):
        if not sheet_name or len(sheet_name) == 0:
            raise InvalidArgumentError("sheet_name is required for xlsx/xlsb files")
        
        for sheet in sheet_name:
            if format == FileFormat.xlsx:
                df, metadata = read_xlsx_sheet(
                    file_path, sheet, include_hidden,
                    max_rows=max_rows, max_cols=max_cols,
                    enable_style_scan=enable_style_scan,
                    enable_border_scan=enable_border_scan,
                    style_scan_limit=style_scan_limit,
                    border_scan_limit=border_scan_limit
                )
            else:  # xlsb
                df, metadata = read_xlsb_sheet(file_path, sheet, include_hidden)
            result[sheet] = (df, metadata)
    
    return result
